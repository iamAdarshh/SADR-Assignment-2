import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the first dataset (the one previously cleaned)
file_path_1 = 'data/GPT/cleaned_combined.xlsx'  # Replace with the actual path of the first dataset
data_1 = pd.read_excel(file_path_1)

# Load the second dataset
file_path_2 = 'data/Analysis_03.12.2024.xlsx'  # Replace with the actual path of the second dataset
data_2 = pd.read_excel(file_path_2)

# Merge the two datasets on 'IssueId', keeping all rows from both datasets
merged_data = pd.merge(data_1, data_2, on="IssueId", how="outer", suffixes=('_left', '_right'))

# Drop duplicate rows
merged_data = merged_data.drop_duplicates()

# Remove duplicate columns by keeping the left side column (e.g., 'column_left' or 'column_right')
# Iterate over columns and merge where suffixes match
for col in merged_data.columns:
    if col.endswith('_left') and col[:-5] + '_right' in merged_data.columns:
        merged_data[col[:-5]] = merged_data[col].combine_first(merged_data[col[:-5] + '_right'])
        merged_data.drop([col, col[:-5] + '_right'], axis=1, inplace=True)

# Save the merged data to a new Excel file
output_file_path = 'data/output/Analysis_07.12.2024.xlsx'  # Replace with desired output path
merged_data.to_excel(output_file_path, index=False)

# Load the workbook and worksheet
wb = load_workbook('data/output/Analysis_07.12.2024.xlsx')  # Path to the saved temporary Excel file
ws = wb.active

# Define the color schemes
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

# Iterate over rows in the sheet (starting from row 2 to avoid header)
for row in ws.iter_rows(min_row=2, min_col=1, max_col=len(merged_data.columns), max_row=len(merged_data) + 1):
    # Extract the values for the relevant columns
    id_val = row[merged_data.columns.get_loc("Id")].value
    issueid_val = row[merged_data.columns.get_loc("IssueId")].value
    filename_val = row[merged_data.columns.get_loc("FileName")].value

    # Apply row color based on the conditions
    if id_val and issueid_val and filename_val:  # Green: all three have values
        for cell in row:
            cell.fill = green_fill
    elif id_val and issueid_val:  # Yellow: only Id and IssueId have values
        for cell in row:
            cell.fill = yellow_fill
    elif issueid_val and filename_val:  # Blue: only IssueId and FileName have values
        for cell in row:
            cell.fill = blue_fill

# Save the updated workbook with colors
output_file_path = 'data/output/Analysis_07.12.2024_colored.xlsx'  # Replace with the final output path
wb.save(output_file_path)

print("The final colored Excel file has been saved to:", output_file_path)


# Output the merged data path
print("Merged data has been saved to:", output_file_path)
